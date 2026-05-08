"""Local-only Excel logger for submitted credit applications.

Each `/assess` call appends one row to `bank_logs.xlsx` in the project root,
capturing the inputs the user provided and the full assessment output
(decision, statistics, financial breakdown, drivers, narrative).

Failures here are swallowed — logging must never break the API response.
"""

from __future__ import annotations

import json
import threading
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .schemas import AssessmentRequest, AssessmentResponse
from .services import compute_request_fingerprint

LOG_PATH = Path(__file__).resolve().parents[1] / "bank_logs.xlsx"
SHEET_NAME = "applications"

_HEADERS: list[str] = [
    "timestamp",
    "request_fingerprint",
    # --- Input: applicant ---
    "bank_type",
    "language",
    "gross_salary",
    "loan_amount",
    "loan_months",
    "employment_type",
    "age",
    # --- Input: SIMAH summary ---
    "simah_total_debt",
    "simah_max_overdue",
    "simah_inquiries_last_month",
    "simah_credit_history_days",
    "simah_max_dpd",
    "simah_raw_features_json",
    # --- Output: decision + statistics ---
    "decision",
    "risk_level",
    "passed_hard_rules",
    "hard_rule_rejection",
    "failed_rules",
    "pd_prob",
    "model_pd",
    "credit_score",
    "dbr",
    "final_dbr",
    "sama_dbr_cap",
    "max_pd_allowed",
    "pd_threshold",
    "min_score",
    # --- Output: per-gate snapshot ---
    "gate_hard_rules",
    "gate_pd_limit",
    "gate_credit_score",
    "gate_final_dbr",
    "gate_profitability",
    # --- Output: financial breakdown ---
    "offered_interest_rate",
    "final_monthly_payment",
    "expected_revenue",
    "expected_loss",
    "expected_profit",
    # --- Output: decision drivers (top SHAP) ---
    "decision_drivers_json",
    # --- Output: narrative ---
    "risk_summary",
    "key_strengths",
    "key_concerns",
    "decision_explanation",
    "suggested_actions",
]

_lock = threading.Lock()


def _bullets(items: list[str] | None) -> str:
    if not items:
        return ""
    return "\n".join(f"- {str(i)}" for i in items)


def _ensure_workbook() -> Workbook:
    if LOG_PATH.exists():
        try:
            wb = load_workbook(LOG_PATH)
        except Exception:
            wb = Workbook()
            wb.remove(wb.active)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        first_row = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if first_row != _HEADERS:
            ws.delete_rows(1, ws.max_row)
            ws.append(_HEADERS)
    else:
        # Drop the leftover default sheet if it's empty so we don't ship clutter.
        for name in list(wb.sheetnames):
            sheet = wb[name]
            if sheet.max_row <= 1 and (sheet.max_column <= 1) and (sheet.cell(1, 1).value in (None, "")):
                wb.remove(sheet)
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(_HEADERS)

    return wb


def _build_row(req: AssessmentRequest, resp: AssessmentResponse) -> list[Any]:
    simah = req.simah_profile or {}
    raw_features = simah.get("raw_features") if isinstance(simah, dict) else None

    drivers = [
        {
            "feature": d.get("feature"),
            "shap_value": d.get("shap_value"),
            "direction": d.get("direction"),
        }
        for d in (resp.shap_top5 or [])
    ]

    # `gate_results` may be missing on legacy responses — degrade gracefully.
    gates = resp.gate_results or {}

    # Narrative is dual-language now; the spreadsheet only logs the language
    # the applicant submitted with so the columns stay narrow.
    lang = req.language if req.language in ("en", "ar") else "en"

    def _pick_text(value: Any) -> str:
        if isinstance(value, dict):
            return str(value.get(lang) or value.get("en") or "")
        return str(value or "")

    def _pick_list(value: Any) -> list[str]:
        if isinstance(value, dict):
            picked = value.get(lang) or value.get("en") or []
            return list(picked) if isinstance(picked, list) else []
        return list(value) if isinstance(value, list) else []

    return [
        datetime.now().isoformat(timespec="seconds"),
        compute_request_fingerprint(req),
        req.bank_type,
        req.language,
        req.gross_salary,
        req.loan_amount,
        req.loan_months,
        req.employment_type,
        req.age,
        simah.get("total_debt") if isinstance(simah, dict) else None,
        simah.get("max_overdue") if isinstance(simah, dict) else None,
        simah.get("inquiries_last_month") if isinstance(simah, dict) else None,
        simah.get("credit_history_days") if isinstance(simah, dict) else None,
        simah.get("max_dpd") if isinstance(simah, dict) else None,
        json.dumps(raw_features, ensure_ascii=False, default=str) if raw_features else "",
        resp.decision,
        resp.risk_level,
        resp.passed_hard_rules,
        _pick_text(resp.hard_rule_rejection),
        ", ".join(resp.failed_rules or []),
        # Pricing/model fields stay None when the model was bypassed — openpyxl
        # writes them as empty cells, which reads as "null" rather than "0".
        resp.pd_prob,
        resp.model_pd,
        resp.credit_score,
        resp.dbr,
        resp.final_dbr,
        resp.sama_dbr_cap,
        resp.max_pd_allowed,
        resp.pd_threshold,
        resp.min_score,
        gates.get("hard_rules"),
        gates.get("pd_limit"),
        gates.get("credit_score"),
        gates.get("final_dbr"),
        gates.get("profitability"),
        resp.offered_interest_rate,
        resp.final_monthly_payment,
        resp.expected_revenue,
        resp.expected_loss,
        resp.expected_profit,
        json.dumps(drivers, ensure_ascii=False, default=str),
        _pick_text(resp.risk_summary),
        _bullets(_pick_list(resp.key_strengths)),
        _bullets(_pick_list(resp.key_concerns)),
        _pick_text(resp.decision_explanation),
        _bullets(_pick_list(resp.suggested_actions)),
    ]


def log_assessment(req: AssessmentRequest, resp: AssessmentResponse) -> None:
    """Append one row for this submission. Never raises."""
    try:
        with _lock:
            wb = _ensure_workbook()
            ws = wb[SHEET_NAME]
            ws.append(_build_row(req, resp))
            wb.save(LOG_PATH)
    except Exception:
        # Local logging is best-effort. If the file is open in Excel or the
        # disk is read-only, the assessment itself must still succeed.
        pass
